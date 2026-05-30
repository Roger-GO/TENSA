"""Integration: clone-on-write edit flow against a real ANDES System (Unit 21).

Per the plan's integration scenario: load kundur_full → init clone → edit a
controller param (TGOV1.T1 — a time-constant param, not the per-unit-normalised
droop ``R``) → reload + setup → observe the new value via
``ss.TGOV1.T1.v[idx]`` → undo → original restored.

Exercises the ``Wrapper`` delegation methods (``init_clone`` /
``apply_clone_edit`` / ``undo_clone_edit`` / ``redo_clone_edit``) — the exact
surface the worker dispatches to — without spinning a server (per the env note
that the acceptance suite collides with the running dev stack).
"""

from __future__ import annotations

import shutil
from pathlib import Path

import pytest

from andes_app.core.wrapper import Wrapper

pytestmark = pytest.mark.integration


def _bundled_cases_dir() -> Path:
    pytest.importorskip("andes")
    import andes

    return Path(andes.__file__).parent / "cases"


@pytest.fixture
def kundur_wrapper(tmp_path: Path) -> Wrapper:
    cases = _bundled_cases_dir()
    workspace = tmp_path / "ws"
    workspace.mkdir()
    case = workspace / "kundur_full.xlsx"
    shutil.copy2(cases / "kundur" / "kundur_full.xlsx", case)
    w = Wrapper(workspace=workspace, session_id="edit-flow")
    w.load_case(case)
    return w


def _t1_first(w: Wrapper) -> float:
    return float(list(w._ss.TGOV1.T1.v)[0])


def test_edit_reload_setup_reflects_new_value(kundur_wrapper: Wrapper) -> None:
    w = kundur_wrapper
    init = w.init_clone()
    assert init["already_initialized"] is False

    payload = w.apply_clone_edit("TGOV1", "1", "T1", 0.6)
    assert payload["new_value"] == pytest.approx(0.6)
    assert payload["undo_depth"] == 1
    # The live System (post setup) reflects the edit.
    assert _t1_first(w) == pytest.approx(0.6)


def test_undo_restores_original(kundur_wrapper: Wrapper) -> None:
    w = kundur_wrapper
    original = _t1_first(w)  # triggers setup via run? no — read pre-edit
    # Re-load to a clean pre-setup state for a fair original read after setup.
    w.init_clone()
    w.apply_clone_edit("TGOV1", "1", "T1", 0.6)
    assert _t1_first(w) == pytest.approx(0.6)

    undo = w.undo_clone_edit()
    assert undo["undo_depth"] == 0
    assert undo["redo_depth"] == 1
    assert _t1_first(w) == pytest.approx(original)


def test_redo_round_trips(kundur_wrapper: Wrapper) -> None:
    w = kundur_wrapper
    w.init_clone()
    w.apply_clone_edit("TGOV1", "1", "T1", 0.6)
    w.undo_clone_edit()
    w.redo_clone_edit()
    assert _t1_first(w) == pytest.approx(0.6)


def test_edit_then_pflow_runs_on_new_system(kundur_wrapper: Wrapper) -> None:
    """After a clone edit the re-setup System is runnable end-to-end."""
    w = kundur_wrapper
    w.init_clone()
    w.apply_clone_edit("TGOV1", "1", "T1", 0.6)
    result = w.run_pflow()
    assert result.converged is True


def test_edit_does_not_touch_original_file(
    kundur_wrapper: Wrapper, tmp_path: Path
) -> None:
    from openpyxl import load_workbook

    w = kundur_wrapper
    case = w._case_path
    assert case is not None
    w.init_clone()
    w.apply_clone_edit("TGOV1", "1", "T1", 0.6)

    wb = load_workbook(str(case), read_only=True)
    rows = list(wb["TGOV1"].iter_rows(values_only=True))
    wb.close()
    t1_col = list(rows[0]).index("T1")
    assert rows[1][t1_col] == pytest.approx(0.49)


def test_reset_clone_reverts_and_deletes_dir(kundur_wrapper: Wrapper) -> None:
    w = kundur_wrapper
    w.init_clone()
    mgr = w._clone_mgr()
    clone_dir = mgr.clone_dir
    w.apply_clone_edit("TGOV1", "1", "T1", 0.6)
    assert _t1_first(w) == pytest.approx(0.6)

    w.reset_clone()
    assert clone_dir is not None
    assert not clone_dir.exists()
    assert _t1_first(w) == pytest.approx(0.49)
