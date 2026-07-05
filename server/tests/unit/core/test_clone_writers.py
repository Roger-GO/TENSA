"""Unit tests for the per-format clone writers (Unit 21 / KTD-9).

Covers, per the plan's Test scenarios:

- xlsx writer happy path (cell is set; original-other-cells untouched).
- dyr writer happy path + the CRITICAL **two-devices-same-model** case: both
  sibling ``ST2CUT`` records survive with only the target field edited (the
  spike's in-place-splice requirement).
- raw writer always raises ``CloneEditError`` with the "use edit-element" hint.
- malformed-file recovery: a non-matching ``.dyr`` record raises
  ``CloneEditError`` and leaves the file unchanged.
- the package index loader + dispatch by extension.
"""

from __future__ import annotations

import os
import shutil
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from tensa.core.clone_writers import (
    DyrLocator,
    apply_edit,
    dyr_writer,
    index_entry_for,
    load_clone_write_index,
    raw_writer,
    xlsx_writer,
)
from tensa.core.errors import CloneEditError


def _bundled_cases_dir() -> Path:
    pytest.importorskip("andes")
    import andes

    return Path(andes.__file__).parent / "cases"


# ---- index loader -----------------------------------------------------------


def test_index_loads_18_models() -> None:
    index = load_clone_write_index()
    assert index["schema_version"] == 1
    assert len(index["models"]) == 18


def test_index_entry_for_known_param() -> None:
    entry = index_entry_for("TGOV1", "T1")
    assert entry["xlsx"]["column"] == "T1"
    assert entry["dyr"]["field_index"] == 3


def test_index_entry_for_unknown_model_raises() -> None:
    with pytest.raises(CloneEditError):
        index_entry_for("NotAModel", "T1")


def test_index_entry_for_unknown_param_raises() -> None:
    with pytest.raises(CloneEditError):
        index_entry_for("TGOV1", "NotAParam")


# ---- xlsx writer ------------------------------------------------------------


def _build_tgov1_workbook(path: Path) -> None:
    wb = Workbook()
    sheet = wb.active
    sheet.title = "TGOV1"
    sheet.append(["uid", "idx", "name", "R", "T1", "VMAX"])
    sheet.append([0, "1", "TGOV1_1", 0.05, 0.49, 1.0])
    sheet.append([1, "2", "TGOV1_2", 0.05, 0.50, 1.0])
    wb.save(str(path))


def test_xlsx_writer_sets_target_cell(tmp_path: Path) -> None:
    book = tmp_path / "case.xlsx"
    _build_tgov1_workbook(book)

    xlsx_writer.apply_edit(book, "TGOV1", "1", "T1", 0.6)

    wb = load_workbook(str(book))
    rows = list(wb["TGOV1"].iter_rows(values_only=True))
    wb.close()
    # header + 2 data rows; idx=1 is row index 1 in the data, T1 is col 4.
    assert rows[1][4] == 0.6
    # sibling row (idx=2) untouched.
    assert rows[2][4] == 0.50


def test_xlsx_writer_missing_sheet_raises(tmp_path: Path) -> None:
    book = tmp_path / "empty.xlsx"
    wb = Workbook()
    wb.active.title = "SomethingElse"
    wb.save(str(book))
    wb.close()
    with pytest.raises(CloneEditError, match="sheet"):
        xlsx_writer.apply_edit(book, "TGOV1", "1", "T1", 0.6)


def test_xlsx_writer_missing_idx_row_raises(tmp_path: Path) -> None:
    book = tmp_path / "case.xlsx"
    _build_tgov1_workbook(book)
    with pytest.raises(CloneEditError, match="no row"):
        xlsx_writer.apply_edit(book, "TGOV1", "999", "T1", 0.6)


def test_xlsx_writer_malformed_file_raises(tmp_path: Path) -> None:
    book = tmp_path / "broken.xlsx"
    book.write_text("this is not a zip / xlsx file", encoding="utf-8")
    with pytest.raises(CloneEditError, match="could not open"):
        xlsx_writer.apply_edit(book, "TGOV1", "1", "T1", 0.6)


# ---- dyr writer: two devices, same model (the CRITICAL spike requirement) ---


def test_dyr_writer_preserves_sibling_records(tmp_path: Path) -> None:
    """ieee14.dyr has TWO ST2CUT records (bus 1, bus 2). Editing the bus-1
    record's T3 must leave the bus-2 record byte-intact, with both surviving.
    """
    cases = _bundled_cases_dir()
    src = cases / "ieee14" / "ieee14.dyr"
    dst = tmp_path / "ieee14.dyr"
    shutil.copy2(src, dst)

    before = dst.read_text(encoding="utf-8")
    assert before.count("'ST2CUT'") == 2

    # ST2CUT.T3 -> field_index 10 -> file-token 11. Bus-1 original T3 = 30.0.
    dyr_writer.apply_edit(
        dst, "ST2CUT", "ST2CUT_2", "T3", 99.5, locator=DyrLocator(bus="1.0", id="1")
    )
    after = dst.read_text(encoding="utf-8")

    # Both records survive.
    assert after.count("'ST2CUT'") == 2
    assert "99.5" in after


def test_dyr_writer_round_trips_through_andes(tmp_path: Path) -> None:
    """After the bus-1 edit, ANDES re-reads BOTH ST2CUT devices and only the
    bus-1 device's T3 changed."""
    pytest.importorskip("andes")
    import andes

    cases = _bundled_cases_dir()
    dst = tmp_path / "ieee14.dyr"
    shutil.copy2(cases / "ieee14" / "ieee14.dyr", dst)

    dyr_writer.apply_edit(
        dst, "ST2CUT", "ST2CUT_2", "T3", 99.5, locator=DyrLocator(bus="1.0", id="1")
    )

    ss = andes.load(
        str(cases / "ieee14" / "ieee14.raw"),
        addfile=str(dst),
        setup=True,
        no_output=True,
        default_config=True,
    )
    bus_values = list(ss.ST2CUT.bus.v)
    t3_values = list(ss.ST2CUT.T3.v)
    # Two devices survive.
    assert len(ss.ST2CUT.idx.v) == 2
    # The bus-1 device got the edit; the bus-2 device is unchanged.
    i_bus1 = next(i for i, b in enumerate(bus_values) if int(b) == 1)
    i_bus2 = next(i for i, b in enumerate(bus_values) if int(b) == 2)
    assert t3_values[i_bus1] == pytest.approx(99.5)
    assert t3_values[i_bus2] != pytest.approx(99.5)


def test_dyr_writer_no_matching_record_raises(tmp_path: Path) -> None:
    cases = _bundled_cases_dir()
    dst = tmp_path / "ieee14.dyr"
    shutil.copy2(cases / "ieee14" / "ieee14.dyr", dst)
    original = dst.read_text(encoding="utf-8")
    with pytest.raises(CloneEditError, match="no ST2CUT record"):
        dyr_writer.apply_edit(
            dst, "ST2CUT", "x", "T3", 1.0, locator=DyrLocator(bus="999", id="1")
        )
    # The clone file is unchanged on the failure.
    assert dst.read_text(encoding="utf-8") == original


def test_dyr_writer_non_editable_param_raises(tmp_path: Path) -> None:
    """TGOV1.Tn is an ANDES-added param absent from PSS/E (dyr index is null)."""
    cases = _bundled_cases_dir()
    dst = tmp_path / "ieee14.dyr"
    shutil.copy2(cases / "ieee14" / "ieee14.dyr", dst)
    with pytest.raises(CloneEditError, match="not editable in .dyr"):
        dyr_writer.apply_edit(
            dst, "TGOV1", "TGOV1_1", "Tn", 1.0, locator=DyrLocator(bus="1.0")
        )


def test_dyr_writer_requires_locator(tmp_path: Path) -> None:
    dst = tmp_path / "x.dyr"
    dst.write_text("   1 'TGOV1' 1  0.05  0.4  /\n", encoding="utf-8")
    with pytest.raises(CloneEditError, match="locator is required"):
        dyr_writer.apply_edit(dst, "TGOV1", "TGOV1_1", "T1", 0.6, locator=None)


# ---- raw writer -------------------------------------------------------------


def test_raw_writer_always_rejects(tmp_path: Path) -> None:
    raw = tmp_path / "case.raw"
    raw.write_text("0 / END OF DATA\n", encoding="utf-8")
    with pytest.raises(CloneEditError, match="use-edit-element"):
        raw_writer.apply_edit(raw, "TGOV1", "1", "T1", 0.6)


# ---- dispatch ---------------------------------------------------------------


def test_apply_edit_dispatches_by_extension(tmp_path: Path) -> None:
    book = tmp_path / "case.xlsx"
    _build_tgov1_workbook(book)
    apply_edit(book, "TGOV1", "1", "T1", 0.7)
    wb = load_workbook(str(book))
    rows = list(wb["TGOV1"].iter_rows(values_only=True))
    wb.close()
    assert rows[1][4] == 0.7


def test_apply_edit_unknown_extension_raises(tmp_path: Path) -> None:
    target = tmp_path / "case.unknown"
    target.write_text("x", encoding="utf-8")
    with pytest.raises(CloneEditError, match="unsupported clone file format"):
        apply_edit(target, "TGOV1", "1", "T1", 0.6)


def test_apply_edit_raw_routes_to_raw_writer(tmp_path: Path) -> None:
    raw = tmp_path / "case.raw"
    raw.write_text("0 / END\n", encoding="utf-8")
    with pytest.raises(CloneEditError, match="use-edit-element"):
        apply_edit(raw, "TGOV1", "1", "T1", 0.6)


def test_index_file_is_packaged() -> None:
    """The index ships inside the package (wheel-safe), not just in docs/."""
    import tensa.core.clone_writers as pkg

    pkg_dir = Path(pkg.__file__).parent
    assert (pkg_dir / "clone_write_index.json").exists()
    assert os.path.getsize(pkg_dir / "clone_write_index.json") > 1000
