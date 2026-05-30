"""Unit tests for ``andes_app.core.clone_manager.CloneManager`` (Unit 21).

Covers, per the plan's Test scenarios:

- init_clone copies the originals to scratch; the originals are never touched.
- edit → reload → read-back reflects the new value; undo restores; redo
  round-trips.
- save_as writes the clone to the workspace under the chosen name.
- stack cap = 50, LRU eviction (oldest evicted; redo cleared by a new edit).
- reset_clone deletes the clone dir + reverts to the originals.
- save-as name validation rejects traversal / separators.

The functional tests use a real ANDES System (kundur_full.xlsx). The
stack-cap test stubs the reload to stay fast across 51 edits.
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook, load_workbook

from andes_app.core.clone_manager import UNDO_STACK_CAP, CloneManager
from andes_app.core.errors import CloneEditError, ElementValidationError
from andes_app.core.wrapper import Wrapper

pytestmark = pytest.mark.integration  # uses a real ANDES System


def _bundled_cases_dir() -> Path:
    pytest.importorskip("andes")
    import andes

    return Path(andes.__file__).parent / "cases"


@pytest.fixture
def workspace_with_kundur(tmp_path: Path) -> tuple[Path, Path]:
    """A tmp workspace holding a private copy of kundur_full.xlsx.

    Returns ``(workspace, case_path)``. Tests never mutate the bundled file.
    """
    cases = _bundled_cases_dir()
    workspace = tmp_path / "ws"
    workspace.mkdir()
    case = workspace / "kundur_full.xlsx"
    shutil.copy2(cases / "kundur" / "kundur_full.xlsx", case)
    return workspace, case


@pytest.fixture
def loaded_wrapper(workspace_with_kundur: tuple[Path, Path]) -> Wrapper:
    workspace, case = workspace_with_kundur
    w = Wrapper(workspace=workspace, session_id="sess-test")
    w.load_case(case)
    return w


# ---- init -------------------------------------------------------------------


def test_init_clone_copies_originals(loaded_wrapper: Wrapper) -> None:
    mgr = loaded_wrapper._clone_mgr()
    result = mgr.init_clone()
    assert result.already_initialized is False
    assert mgr.is_initialized
    assert mgr.clone_dir is not None
    assert mgr.clone_dir.exists()
    assert len(mgr.clone_paths) == 1
    assert mgr.clone_paths[0].name == "kundur_full.xlsx"
    # The clone lives under <workspace>/.sessions/<session_id>/clone/.
    assert ".sessions" in str(mgr.clone_dir)
    assert "sess-test" in str(mgr.clone_dir)


def test_init_clone_is_idempotent(loaded_wrapper: Wrapper) -> None:
    mgr = loaded_wrapper._clone_mgr()
    mgr.init_clone()
    second = mgr.init_clone()
    assert second.already_initialized is True


def test_init_clone_blank_session_rejected(tmp_path: Path) -> None:
    w = Wrapper(workspace=tmp_path, session_id="s")
    w.create_blank()
    from andes_app.core.errors import NoCaseLoadedError

    with pytest.raises(NoCaseLoadedError):
        w.init_clone()


# ---- edit + undo + redo -----------------------------------------------------


def test_edit_updates_value_and_leaves_original_untouched(
    loaded_wrapper: Wrapper, workspace_with_kundur: tuple[Path, Path]
) -> None:
    _, case = workspace_with_kundur
    mgr = loaded_wrapper._clone_mgr()
    result = mgr.apply_edit("TGOV1", "1", "T1", 0.6)
    assert result.new_value == pytest.approx(0.6)
    assert result.undo_depth == 1
    assert result.redo_depth == 0
    # Original file is untouched (T1 col).
    wb = load_workbook(str(case), read_only=True)
    rows = list(wb["TGOV1"].iter_rows(values_only=True))
    wb.close()
    header = list(rows[0])
    t1_col = header.index("T1")
    assert rows[1][t1_col] == pytest.approx(0.49)


def test_undo_restores_original_value(loaded_wrapper: Wrapper) -> None:
    mgr = loaded_wrapper._clone_mgr()
    mgr.apply_edit("TGOV1", "1", "T1", 0.6)
    undo = mgr.undo()
    assert undo.undo_depth == 0
    assert undo.redo_depth == 1
    assert list(loaded_wrapper._ss.TGOV1.T1.v)[0] == pytest.approx(0.49)


def test_redo_reapplies_edit(loaded_wrapper: Wrapper) -> None:
    mgr = loaded_wrapper._clone_mgr()
    mgr.apply_edit("TGOV1", "1", "T1", 0.6)
    mgr.undo()
    redo = mgr.redo()
    assert redo.undo_depth == 1
    assert redo.redo_depth == 0
    assert list(loaded_wrapper._ss.TGOV1.T1.v)[0] == pytest.approx(0.6)


def test_new_edit_clears_redo_stack(loaded_wrapper: Wrapper) -> None:
    mgr = loaded_wrapper._clone_mgr()
    mgr.apply_edit("TGOV1", "1", "T1", 0.6)
    mgr.undo()
    assert len(mgr.redo_stack) == 1
    # A NEW edit after an undo clears the redo stack (KTD-10).
    result = mgr.apply_edit("TGOV1", "1", "T1", 0.7)
    assert result.redo_depth == 0
    assert len(mgr.redo_stack) == 0


def test_undo_with_empty_stack_raises(loaded_wrapper: Wrapper) -> None:
    mgr = loaded_wrapper._clone_mgr()
    mgr.init_clone()
    with pytest.raises(CloneEditError, match="nothing to undo"):
        mgr.undo()


def test_redo_with_empty_stack_raises(loaded_wrapper: Wrapper) -> None:
    mgr = loaded_wrapper._clone_mgr()
    mgr.init_clone()
    with pytest.raises(CloneEditError, match="nothing to redo"):
        mgr.redo()


# ---- whitelist (defence in depth) -------------------------------------------


def test_edit_non_controller_model_rejected(loaded_wrapper: Wrapper) -> None:
    mgr = loaded_wrapper._clone_mgr()
    with pytest.raises(ElementValidationError):
        mgr.apply_edit("Bus", "1", "Vn", 1.0)


def test_edit_unknown_param_rejected(loaded_wrapper: Wrapper) -> None:
    mgr = loaded_wrapper._clone_mgr()
    with pytest.raises(ElementValidationError):
        mgr.apply_edit("TGOV1", "1", "NotAParam", 1.0)


# ---- save-as ----------------------------------------------------------------


def test_save_as_writes_to_workspace(
    loaded_wrapper: Wrapper, workspace_with_kundur: tuple[Path, Path]
) -> None:
    workspace, _ = workspace_with_kundur
    mgr = loaded_wrapper._clone_mgr()
    mgr.apply_edit("TGOV1", "1", "T1", 0.6)
    result = mgr.save_as("kundur_tuned")
    assert result.name == "kundur_tuned"
    saved = workspace / "kundur_tuned.xlsx"
    assert saved.exists()
    wb = load_workbook(str(saved), read_only=True)
    rows = list(wb["TGOV1"].iter_rows(values_only=True))
    wb.close()
    t1_col = list(rows[0]).index("T1")
    assert rows[1][t1_col] == pytest.approx(0.6)


@pytest.mark.parametrize("bad", ["../escape", "a/b", "with space", ".hidden", ""])
def test_save_as_rejects_unsafe_names(loaded_wrapper: Wrapper, bad: str) -> None:
    mgr = loaded_wrapper._clone_mgr()
    mgr.init_clone()
    with pytest.raises(CloneEditError):
        mgr.save_as(bad)


def test_save_as_without_clone_raises(loaded_wrapper: Wrapper) -> None:
    mgr = loaded_wrapper._clone_mgr()
    with pytest.raises(CloneEditError, match="no clone"):
        mgr.save_as("name")


# ---- reset ------------------------------------------------------------------


def test_reset_deletes_clone_dir_and_reverts(loaded_wrapper: Wrapper) -> None:
    mgr = loaded_wrapper._clone_mgr()
    mgr.apply_edit("TGOV1", "1", "T1", 0.6)
    clone_dir = mgr.clone_dir
    assert clone_dir is not None and clone_dir.exists()
    mgr.reset_clone()
    assert not clone_dir.exists()
    assert mgr.is_initialized is False
    assert mgr.clone_paths == []
    # Live System reverted to the original value.
    assert list(loaded_wrapper._ss.TGOV1.T1.v)[0] == pytest.approx(0.49)


def test_reset_then_edit_reinitialises(loaded_wrapper: Wrapper) -> None:
    mgr = loaded_wrapper._clone_mgr()
    mgr.apply_edit("TGOV1", "1", "T1", 0.6)
    mgr.reset_clone()
    result = mgr.apply_edit("TGOV1", "1", "T1", 0.55)
    assert mgr.is_initialized
    assert result.new_value == pytest.approx(0.55)


# ---- stack cap / LRU eviction (fast: reload stubbed) ------------------------


def test_stack_cap_lru_eviction(tmp_path: Path, monkeypatch: Any) -> None:
    """The 51st edit evicts the oldest stack entry (cap = 50)."""
    assert UNDO_STACK_CAP == 50

    # Real tmp workbook so the xlsx writer succeeds; reload is stubbed so the
    # 51 edits stay fast (no andes.load per edit).
    book = tmp_path / "case.xlsx"
    wb = Workbook()
    sheet = wb.active
    sheet.title = "TGOV1"
    sheet.append(["uid", "idx", "name", "T1"])
    sheet.append([0, "1", "TGOV1_1", 0.49])
    wb.save(str(book))
    wb.close()

    class _StubWrapper:
        _case_path = book
        _addfiles: list[Path] = []
        _ss = None

        def reload_case(self) -> None:  # pragma: no cover — not hit here
            pass

    mgr = CloneManager(
        wrapper=_StubWrapper(),  # type: ignore[arg-type]
        workspace=tmp_path,
        session_id="cap-test",
    )
    monkeypatch.setattr(mgr, "_reload_from_clone", lambda: None)
    monkeypatch.setattr(
        mgr, "_read_back_value", lambda model, idx, param: None
    )
    # Skip the dyr locator path — target is xlsx, so no System needed.
    monkeypatch.setattr(mgr, "_dyr_locator", lambda model, idx: None)

    for n in range(UNDO_STACK_CAP + 1):
        mgr.apply_edit("TGOV1", "1", "T1", 0.4 + n * 0.001)

    # 51 edits, cap 50 → exactly 50 retained, oldest evicted.
    assert len(mgr.undo_stack) == UNDO_STACK_CAP
