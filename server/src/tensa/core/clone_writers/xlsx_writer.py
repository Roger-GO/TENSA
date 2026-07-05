"""ANDES-native ``.xlsx`` clone writer (Unit 21 / KTD-9).

Edits a single ``(model, idx, param)`` cell in place via openpyxl:

1. Open the workbook; select the sheet named ``model``.
2. Read the header row; find the ``idx`` column and the ``param`` column.
3. Find the row whose ``idx`` cell equals ``idx``; set its ``param`` cell.

Every editable controller param is ``.xlsx``-editable (the ANDES exporter
writes a column per param — 100% coverage per the spike). The sheet / idx
column / param column are all read from the spike index, never hardcoded.

Pattern mirrors ``api/routes/profiles.py``'s openpyxl usage.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from tensa.core.clone_writers import index_entry_for, load_clone_write_index
from tensa.core.errors import CloneEditError


def apply_edit(
    file_path: str | Path,
    model: str,
    idx: str,
    param: str,
    value: Any,
) -> None:
    """Set ``model.param`` for device ``idx`` in the ``.xlsx`` clone in place.

    Raises :class:`CloneEditError` when the sheet, idx column, idx row, or
    param column is absent — i.e. a malformed / hand-edited workbook or a
    drift between the live whitelist and the spike index.
    """
    path = Path(file_path)
    index = load_clone_write_index()
    model_entry = index.get("models", {}).get(model)
    if not isinstance(model_entry, dict):
        raise CloneEditError(
            f"model {model!r} is not in the clone-write index"
        )
    xlsx_meta = model_entry.get("xlsx")
    if not isinstance(xlsx_meta, dict):
        raise CloneEditError(
            f"model {model!r} has no .xlsx edit metadata in the clone-write index"
        )
    sheet_name = str(xlsx_meta.get("sheet", model))
    idx_column_name = str(xlsx_meta.get("idx_column", "idx"))

    param_entry = index_entry_for(model, param)
    xlsx_param = param_entry.get("xlsx")
    if not isinstance(xlsx_param, dict) or "column" not in xlsx_param:
        raise CloneEditError(
            f"param {param!r} on model {model!r} has no .xlsx column in the "
            "clone-write index"
        )
    param_column_name = str(xlsx_param["column"])

    try:
        workbook = load_workbook(filename=str(path))
    except Exception as exc:  # noqa: BLE001 — wrap any openpyxl/zip error
        raise CloneEditError(
            f"could not open clone workbook {path.name!r}: {exc}"
        ) from exc

    try:
        if sheet_name not in workbook.sheetnames:
            raise CloneEditError(
                f"sheet {sheet_name!r} not found in clone workbook {path.name!r}"
            )
        sheet = workbook[sheet_name]

        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header is None:
            raise CloneEditError(
                f"sheet {sheet_name!r} in {path.name!r} has no header row"
            )
        header_list = [str(h) if h is not None else "" for h in header]
        try:
            idx_col = header_list.index(idx_column_name)
        except ValueError as exc:
            raise CloneEditError(
                f"idx column {idx_column_name!r} not found in sheet "
                f"{sheet_name!r} header"
            ) from exc
        try:
            param_col = header_list.index(param_column_name)
        except ValueError as exc:
            raise CloneEditError(
                f"param column {param_column_name!r} not found in sheet "
                f"{sheet_name!r} header"
            ) from exc

        idx_str = str(idx)
        target_row: int | None = None
        # Rows are 1-based; data starts at row 2 (row 1 is the header). The
        # cell column is 1-based, so the 0-based ``idx_col`` maps to column
        # ``idx_col + 1``.
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=idx_col + 1).value
            if cell_value is not None and str(cell_value) == idx_str:
                target_row = row
                break
        if target_row is None:
            raise CloneEditError(
                f"no row with {idx_column_name}={idx!r} in sheet {sheet_name!r}"
            )

        sheet.cell(row=target_row, column=param_col + 1).value = value
        workbook.save(str(path))
    finally:
        workbook.close()
